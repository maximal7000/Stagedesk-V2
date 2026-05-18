"""
Django Ninja API für Haushalts-Management
GLOBAL: Alle Benutzer teilen sich die gleichen Haushalte
"""
from typing import List, Optional
from ninja import Router, Schema, File
from ninja.files import UploadedFile
from ninja.errors import HttpError
from django.shortcuts import get_object_or_404
from django.http import HttpRequest
from django.db.models import Max

from core.auth import keycloak_auth
from core.audit import log as audit_log
from users.api import is_admin
from users.models import UserProfile
from .models import Haushalt, Artikel, Kategorie, SammelQuittung
from .schemas import (
    HaushaltSchema,
    HaushaltCreateSchema,
    HaushaltUpdateSchema,
    ArtikelSchema,
    ArtikelCreateSchema,
    ArtikelUpdateSchema,
    ArtikelReorderSchema,
    KategorieSchema,
    KategorieCreateSchema,
    LinkParseRequestSchema,
    LinkParseResponseSchema,
)
from .services import LinkParserService

# Router für Haushalte
haushalte_router = Router(tags=["Haushalte"])


def get_user_id_from_token(request: HttpRequest) -> str:
    """Extrahiert die Benutzer-ID aus dem Token (für Logging/Audit)"""
    if hasattr(request, 'auth') and request.auth:
        return request.auth.get('preferred_username') or request.auth.get('sub')
    return 'anonymous'


def require_perm(request, code: str):
    """Erlaubt Zugriff nur wenn der User Admin ist oder die Permission besitzt."""
    if is_admin(request):
        return
    kid = request.auth.get('sub', '')
    try:
        profile = UserProfile.objects.get(keycloak_id=kid)
        if profile.has_permission(code, False):
            return
    except UserProfile.DoesNotExist:
        pass
    raise HttpError(403, "Keine Berechtigung")


# ==================== Haushalt Endpoints (GLOBAL MIT AUTH) ====================

@haushalte_router.get("/", response=List[HaushaltSchema], auth=keycloak_auth)
def list_haushalte(request):
    """Alle Haushalte auflisten (global für alle authentifizierten Benutzer)"""
    require_perm(request, 'haushalte.view')
    haushalte = Haushalt.objects.all()
    return haushalte


@haushalte_router.get("/{haushalt_id}", response=HaushaltSchema, auth=keycloak_auth)
def get_haushalt(request, haushalt_id: int):
    """Einzelnen Haushalt abrufen"""
    require_perm(request, 'haushalte.view')
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    return haushalt


@haushalte_router.post("/", response=HaushaltSchema, auth=keycloak_auth)
def create_haushalt(request, payload: HaushaltCreateSchema):
    """Neuen Haushalt erstellen"""
    require_perm(request, 'haushalte.create')
    user_id = get_user_id_from_token(request)
    haushalt = Haushalt.objects.create(
        name=payload.name,
        beschreibung=payload.beschreibung,
        budget_konsumitiv=payload.budget_konsumitiv,
        budget_investiv=payload.budget_investiv,
        benutzer_id=user_id,  # Wer hat erstellt (für Audit)
    )
    audit_log(request, 'erstellt', 'haushalt', haushalt.id, haushalt.name)
    return haushalt


@haushalte_router.put("/{haushalt_id}", response=HaushaltSchema, auth=keycloak_auth)
def update_haushalt(request, haushalt_id: int, payload: HaushaltUpdateSchema):
    """Haushalt aktualisieren"""
    require_perm(request, 'haushalte.edit')
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)

    if payload.name is not None:
        haushalt.name = payload.name
    if payload.beschreibung is not None:
        haushalt.beschreibung = payload.beschreibung
    if payload.budget_konsumitiv is not None:
        haushalt.budget_konsumitiv = payload.budget_konsumitiv
    if payload.budget_investiv is not None:
        haushalt.budget_investiv = payload.budget_investiv

    haushalt.save()
    audit_log(request, 'aktualisiert', 'haushalt', haushalt.id, haushalt.name)
    return haushalt


@haushalte_router.delete("/{haushalt_id}", auth=keycloak_auth)
def delete_haushalt(request, haushalt_id: int):
    """Haushalt löschen"""
    require_perm(request, 'haushalte.delete')
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    name = haushalt.name
    hid = haushalt.id
    haushalt.delete()
    audit_log(request, 'geloescht', 'haushalt', hid, name)
    return {"success": True}


# ==================== Artikel Endpoints (GLOBAL MIT AUTH) ====================

@haushalte_router.get("/{haushalt_id}/artikel", response=List[ArtikelSchema], auth=keycloak_auth)
def list_artikel(request, haushalt_id: int):
    """Alle Artikel eines Haushalts auflisten"""
    require_perm(request, 'haushalte.view')
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    artikel = haushalt.artikel.all()
    return artikel


@haushalte_router.post("/{haushalt_id}/artikel", response=ArtikelSchema, auth=keycloak_auth)
def create_artikel(request, haushalt_id: int, payload: ArtikelCreateSchema):
    """Neuen Artikel zu einem Haushalt hinzufügen"""
    require_perm(request, 'haushalte.edit')
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)

    # Neue Artikel landen ganz unten in der Sortier-Reihenfolge.
    max_sort = haushalt.artikel.aggregate(s=Max('sortierung'))['s'] or 0

    artikel = Artikel.objects.create(
        haushalt=haushalt,
        name=payload.name,
        beschreibung=payload.beschreibung or '',
        preis=payload.preis,
        anzahl=payload.anzahl,
        kategorie=payload.kategorie,
        link=payload.link or '',
        bild_url=payload.bild_url or '',
        status=payload.status or 'beantragt',
        sortierung=max_sort + 1,
    )
    return artikel


@haushalte_router.put("/{haushalt_id}/artikel/reorder", auth=keycloak_auth)
def reorder_artikel(request, haushalt_id: int, payload: ArtikelReorderSchema):
    """Setzt sortierung gemäß übergebener ID-Reihenfolge (für Drag&Drop / ↑↓)."""
    require_perm(request, 'haushalte.edit')
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    for idx, aid in enumerate(payload.ids):
        Artikel.objects.filter(id=aid, haushalt=haushalt).update(sortierung=idx)
    return {"status": "ok", "count": len(payload.ids)}


class ArtikelBulkUpdateSchema(Schema):
    ids: List[int]
    status: Optional[str] = None
    kategorie: Optional[str] = None


@haushalte_router.put("/{haushalt_id}/artikel/bulk", auth=keycloak_auth)
def bulk_update_artikel(request, haushalt_id: int, payload: ArtikelBulkUpdateSchema):
    """Aktualisiert mehrere Artikel auf einmal (Status / Kategorie)."""
    require_perm(request, 'haushalte.edit')
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    qs = Artikel.objects.filter(id__in=payload.ids, haushalt=haushalt)
    update_fields = {}
    if payload.status:
        update_fields['status'] = payload.status
    if payload.kategorie:
        update_fields['kategorie'] = payload.kategorie
    if not update_fields:
        return {"updated": 0}
    count = qs.update(**update_fields)
    audit_log(request, 'aktualisiert', 'haushalt_bulk', haushalt.id,
              f"{count} Artikel: {', '.join(f'{k}={v}' for k, v in update_fields.items())}")
    return {"updated": count}


@haushalte_router.get("/{haushalt_id}/artikel.csv", auth=keycloak_auth)
def export_artikel_csv(request, haushalt_id: int):
    """Exportiert alle Artikel eines Haushalts als CSV."""
    require_perm(request, 'haushalte.view')
    from django.http import HttpResponse
    import csv, io
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=';')
    w.writerow(['Status', 'Kategorie', 'Name', 'Anzahl', 'Preis', 'Gesamt',
                'Link', 'Beschreibung', 'Erstellt', 'Gekauft'])
    for a in haushalt.artikel.all().order_by('sortierung', '-erstellt_am'):
        w.writerow([
            a.get_status_display(),
            a.get_kategorie_display(),
            a.name,
            a.anzahl,
            f'{a.preis:.2f}',
            f'{a.gesamtpreis:.2f}',
            a.link or '',
            (a.beschreibung or '').replace('\n', ' '),
            a.erstellt_am.strftime('%Y-%m-%d %H:%M') if a.erstellt_am else '',
            a.gekauft_am.strftime('%Y-%m-%d') if a.gekauft_am else '',
        ])
    resp = HttpResponse(buf.getvalue().encode('utf-8-sig'), content_type='text/csv; charset=utf-8')
    safe_name = ''.join(c for c in haushalt.name if c.isalnum() or c in '-_ ').strip()
    resp['Content-Disposition'] = f'attachment; filename="{safe_name}_artikel.csv"'
    return resp


@haushalte_router.get("/{haushalt_id}/artikel.xlsx", auth=keycloak_auth)
def export_artikel_xlsx(request, haushalt_id: int):
    """Artikel als Excel-Datei."""
    require_perm(request, 'haushalte.view')
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (haushalt.name or 'Artikel')[:31]

    headers = ['Status', 'Kategorie', 'Name', 'Anzahl', 'Preis €', 'Gesamt €',
               'Link', 'Beschreibung', 'Erstellt', 'Gekauft']
    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill; c.alignment = Alignment(vertical='center')

    sum_total = 0
    for a in haushalt.artikel.all().order_by('sortierung', '-erstellt_am'):
        ws.append([
            a.get_status_display(),
            a.get_kategorie_display(),
            a.name,
            a.anzahl,
            float(a.preis),
            float(a.gesamtpreis),
            a.link or '',
            (a.beschreibung or '').replace('\n', ' '),
            a.erstellt_am.strftime('%Y-%m-%d %H:%M') if a.erstellt_am else '',
            a.gekauft_am.strftime('%Y-%m-%d') if a.gekauft_am else '',
        ])
        sum_total += float(a.gesamtpreis)

    # Spaltenbreiten grob
    widths = [12, 12, 36, 8, 10, 10, 30, 40, 18, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    # Preis-Zellen formatieren
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = '#,##0.00 €'
    # Summenzeile
    ws.append([])
    summary_row = ['', '', '', 'Gesamt:', '', sum_total]
    ws.append(summary_row)
    last = ws.max_row
    ws.cell(row=last, column=4).font = Font(bold=True)
    ws.cell(row=last, column=6).font = Font(bold=True)
    ws.cell(row=last, column=6).number_format = '#,##0.00 €'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = ''.join(c for c in haushalt.name if c.isalnum() or c in '-_ ').strip() or 'haushalt'
    resp = HttpResponse(buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{safe}.xlsx"'
    return resp


@haushalte_router.get("/{haushalt_id}/artikel.pdf", auth=keycloak_auth)
def export_artikel_pdf(request, haushalt_id: int):
    """Artikel als PDF-Bericht."""
    require_perm(request, 'haushalte.view')
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    import io

    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"<b>{haushalt.name}</b>", styles['Title']),
    ]
    if haushalt.beschreibung:
        elements.append(Paragraph(haushalt.beschreibung, styles['Normal']))
    elements.append(Spacer(1, 12))

    data = [['Status', 'Kategorie', 'Name', 'Anz.', 'Preis', 'Gesamt', 'Gekauft']]
    sum_total = 0
    for a in haushalt.artikel.all().order_by('sortierung', '-erstellt_am'):
        data.append([
            a.get_status_display(),
            a.get_kategorie_display(),
            (a.name[:50] + '…') if len(a.name) > 51 else a.name,
            str(a.anzahl),
            f'{float(a.preis):.2f} €',
            f'{float(a.gesamtpreis):.2f} €',
            a.gekauft_am.strftime('%d.%m.%Y') if a.gekauft_am else '',
        ])
        sum_total += float(a.gesamtpreis)
    data.append(['', '', '', '', 'Gesamt:', f'{sum_total:.2f} €', ''])

    tbl = Table(data, repeatRows=1, colWidths=[70, 80, 220, 30, 60, 70, 60])
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#9CA3AF')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F3F4F6')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (4, 1), (5, -1), 'RIGHT'),
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),
    ]))
    elements.append(tbl)
    doc.build(elements)
    buf.seek(0)
    safe = ''.join(c for c in haushalt.name if c.isalnum() or c in '-_ ').strip() or 'haushalt'
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{safe}.pdf"'
    return resp


@haushalte_router.get("/{haushalt_id}/status-summary", auth=keycloak_auth)
def haushalt_status_summary(request, haushalt_id: int):
    """Budget-Aufteilung pro Artikel-Status: zeigt wie viel beantragt,
    bestellt, geliefert etc. ist — pro Kategorie."""
    require_perm(request, 'haushalte.view')
    from django.db.models import Sum, F, DecimalField
    from django.db.models.functions import Coalesce
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    out = {'konsumitiv': {}, 'investiv': {}}
    rows = (haushalt.artikel
            .values('kategorie', 'status')
            .annotate(summe=Coalesce(Sum(F('preis') * F('anzahl'),
                                          output_field=DecimalField(max_digits=12, decimal_places=2)), 0))
            .annotate(anzahl=Sum('anzahl')))
    for r in rows:
        out[r['kategorie']][r['status']] = {
            'summe': float(r['summe'] or 0),
            'anzahl': int(r['anzahl'] or 0),
        }
    return out


@haushalte_router.put("/{haushalt_id}/artikel/{artikel_id}", response=ArtikelSchema, auth=keycloak_auth)
def update_artikel(request, haushalt_id: int, artikel_id: int, payload: ArtikelUpdateSchema):
    """Artikel aktualisieren — loggt Status-Wechsel und größere Änderungen ins Audit."""
    require_perm(request, 'haushalte.edit')
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    artikel = get_object_or_404(Artikel, id=artikel_id, haushalt=haushalt)

    before = {
        'status': artikel.status,
        'kategorie': artikel.kategorie,
        'name': artikel.name,
        'preis': str(artikel.preis),
        'anzahl': artikel.anzahl,
    }
    data = payload.dict(exclude_unset=True)
    for attr, value in data.items():
        setattr(artikel, attr, value)
    artikel.save()

    # Audit-Eintrag pro relevantes Feld, damit ein User-Verlauf entsteht
    relevant = {k: v for k, v in data.items()
                if k in ('status', 'kategorie', 'name', 'preis', 'anzahl', 'gekauft_am')}
    if relevant:
        for field, new_val in relevant.items():
            old_val = before.get(field)
            if str(old_val) == str(new_val):
                continue
            aktion = 'status_geaendert' if field == 'status' else 'aktualisiert'
            audit_log(request, aktion, 'artikel', artikel.id,
                      f'{artikel.name} · {field}: {old_val} → {new_val}',
                      {'feld': field, 'alt': old_val, 'neu': str(new_val)})
    return artikel


# ─── Kommentare am Artikel ────────────────────────────────────────

class KommentarCreateSchema(Schema):
    text: str


@haushalte_router.get("/{haushalt_id}/artikel/{artikel_id}/kommentare", auth=keycloak_auth)
def list_kommentare(request, haushalt_id: int, artikel_id: int):
    require_perm(request, 'haushalte.view')
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    artikel = get_object_or_404(Artikel, id=artikel_id, haushalt=haushalt)
    return [
        {
            'id': k.id,
            'text': k.text,
            'user_username': k.user_username,
            'erstellt_am': k.erstellt_am.isoformat() if k.erstellt_am else None,
        }
        for k in artikel.kommentare.all()
    ]


@haushalte_router.post("/{haushalt_id}/artikel/{artikel_id}/kommentare", auth=keycloak_auth)
def add_kommentar(request, haushalt_id: int, artikel_id: int, payload: KommentarCreateSchema):
    require_perm(request, 'haushalte.view')
    from .models import ArtikelKommentar
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    artikel = get_object_or_404(Artikel, id=artikel_id, haushalt=haushalt)
    k = ArtikelKommentar.objects.create(
        artikel=artikel,
        user_keycloak_id=request.auth.get('sub', '') if request.auth else '',
        user_username=request.auth.get('preferred_username', '') if request.auth else '',
        text=payload.text[:5000],
    )
    return {
        'id': k.id, 'text': k.text, 'user_username': k.user_username,
        'erstellt_am': k.erstellt_am.isoformat(),
    }


@haushalte_router.delete("/{haushalt_id}/artikel/{artikel_id}/kommentare/{kid}", auth=keycloak_auth)
def delete_kommentar(request, haushalt_id: int, artikel_id: int, kid: int):
    from .models import ArtikelKommentar
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    artikel = get_object_or_404(Artikel, id=artikel_id, haushalt=haushalt)
    k = get_object_or_404(ArtikelKommentar, id=kid, artikel=artikel)
    # Eigener Kommentar oder edit-Permission
    own = k.user_keycloak_id and request.auth and k.user_keycloak_id == request.auth.get('sub', '')
    if not own:
        require_perm(request, 'haushalte.edit')
    k.delete()
    return {"status": "deleted"}


# ─── Verlauf eines Artikels (aus Audit-Log) ──────────────────────

@haushalte_router.get("/{haushalt_id}/artikel/{artikel_id}/verlauf", auth=keycloak_auth)
def artikel_verlauf(request, haushalt_id: int, artikel_id: int):
    require_perm(request, 'haushalte.view')
    from inventar.models import AuditLog
    qs = AuditLog.objects.filter(entity_type='artikel', entity_id=artikel_id).order_by('-timestamp')[:200]
    return [
        {
            'id': e.id,
            'aktion': e.aktion,
            'aktion_display': e.get_aktion_display(),
            'details': e.details,
            'entity_name': e.entity_name,
            'user_username': e.user_username,
            'timestamp': e.timestamp.isoformat() if e.timestamp else None,
        } for e in qs
    ]


# ─── Quittungs-Upload ─────────────────────────────────────────────

@haushalte_router.post("/{haushalt_id}/artikel/{artikel_id}/quittung", auth=keycloak_auth)
def upload_quittung(request, haushalt_id: int, artikel_id: int,
                    datei: UploadedFile = File(...)):
    require_perm(request, 'haushalte.edit')
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    artikel = get_object_or_404(Artikel, id=artikel_id, haushalt=haushalt)
    # Alte Datei löschen falls vorhanden
    if artikel.quittung:
        try: artikel.quittung.delete(save=False)
        except Exception: pass
    artikel.quittung = datei
    artikel.save(update_fields=['quittung'])
    audit_log(request, 'erstellt', 'quittung', artikel.id,
              f'{artikel.name} · Quittung hochgeladen')
    return {"quittung_url": artikel.quittung.url if artikel.quittung else None}


# ─── Sammel-Quittungen (eine Quittung an mehrere Artikel) ────────

@haushalte_router.get("/{haushalt_id}/sammelquittungen", auth=keycloak_auth)
def list_sammelquittungen(request, haushalt_id: int):
    require_perm(request, 'haushalte.view')
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    return [{
        "id": q.id,
        "name": q.name or f"Quittung #{q.id}",
        "datei_url": q.datei.url if q.datei else None,
        "hochgeladen_am": q.hochgeladen_am.isoformat(),
        "artikel_ids": list(q.artikel.values_list('id', flat=True)),
    } for q in haushalt.sammelquittungen.prefetch_related('artikel').all()]


class SammelQuittungArtikelSchema(Schema):
    artikel_ids: List[int]


@haushalte_router.post("/{haushalt_id}/sammelquittungen", auth=keycloak_auth)
def upload_sammelquittung(request, haushalt_id: int,
                          datei: UploadedFile = File(...),
                          name: str = '',
                          artikel_ids: str = ''):
    """Lädt eine Quittung hoch und verknüpft sie mit mehreren Artikeln.
    artikel_ids als komma-separierte Liste, z.B. '12,17,23'."""
    require_perm(request, 'haushalte.edit')
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    kid = request.auth.get('sub', '') if request.auth else ''
    q = SammelQuittung.objects.create(
        haushalt=haushalt, name=name[:200], datei=datei, hochgeladen_von=kid,
    )
    ids = [int(x) for x in artikel_ids.split(',') if x.strip().isdigit()]
    if ids:
        artikel = list(Artikel.objects.filter(id__in=ids, haushalt=haushalt))
        for a in artikel:
            a.sammelquittungen.add(q)
    audit_log(request, 'erstellt', 'sammelquittung', q.id,
              f"{haushalt.name} · {len(ids)} Artikel verknüpft")
    return {"id": q.id, "datei_url": q.datei.url, "artikel_anzahl": len(ids)}


@haushalte_router.put("/{haushalt_id}/sammelquittungen/{quittung_id}/artikel", auth=keycloak_auth)
def update_sammelquittung_artikel(request, haushalt_id: int, quittung_id: int,
                                  payload: SammelQuittungArtikelSchema):
    """Ersetzt die Artikel-Verknüpfungen einer Sammel-Quittung."""
    require_perm(request, 'haushalte.edit')
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    q = get_object_or_404(SammelQuittung, id=quittung_id, haushalt=haushalt)
    # Alle alten Verknüpfungen entfernen
    for a in q.artikel.all():
        a.sammelquittungen.remove(q)
    # Neue setzen
    artikel = Artikel.objects.filter(id__in=payload.artikel_ids, haushalt=haushalt)
    for a in artikel:
        a.sammelquittungen.add(q)
    return {"artikel_anzahl": artikel.count()}


@haushalte_router.delete("/{haushalt_id}/sammelquittungen/{quittung_id}", auth=keycloak_auth)
def delete_sammelquittung(request, haushalt_id: int, quittung_id: int):
    require_perm(request, 'haushalte.edit')
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    q = get_object_or_404(SammelQuittung, id=quittung_id, haushalt=haushalt)
    if q.datei:
        try: q.datei.delete(save=False)
        except Exception: pass
    q.delete()
    return {"status": "deleted"}


@haushalte_router.delete("/{haushalt_id}/artikel/{artikel_id}/quittung", auth=keycloak_auth)
def delete_quittung(request, haushalt_id: int, artikel_id: int):
    require_perm(request, 'haushalte.edit')
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    artikel = get_object_or_404(Artikel, id=artikel_id, haushalt=haushalt)
    if artikel.quittung:
        try: artikel.quittung.delete(save=False)
        except Exception: pass
    artikel.quittung = None
    artikel.save(update_fields=['quittung'])
    return {"status": "deleted"}


@haushalte_router.delete("/{haushalt_id}/artikel/{artikel_id}", auth=keycloak_auth)
def delete_artikel(request, haushalt_id: int, artikel_id: int):
    """Artikel löschen"""
    require_perm(request, 'haushalte.delete')
    haushalt = get_object_or_404(Haushalt, id=haushalt_id)
    artikel = get_object_or_404(Artikel, id=artikel_id, haushalt=haushalt)
    artikel.delete()
    return {"success": True}


# ==================== Kategorien Endpoints ====================

kategorien_router = Router(tags=["Kategorien"])


@kategorien_router.get("/", response=List[KategorieSchema])
def list_kategorien(request):
    """Alle verfügbaren Kategorien auflisten (öffentlich)"""
    kategorien = Kategorie.objects.all()
    return kategorien


@kategorien_router.post("/", response=KategorieSchema, auth=keycloak_auth)
def create_kategorie(request, payload: KategorieCreateSchema):
    """Neue Kategorie erstellen"""
    require_perm(request, 'haushalte.edit')
    kategorie = Kategorie.objects.create(
        name=payload.name,
        beschreibung=payload.beschreibung,
        icon=payload.icon,
        farbe=payload.farbe,
    )
    return kategorie


# ==================== Link Parser Endpoint ====================

@haushalte_router.post("/parse-link/", response=LinkParseResponseSchema, auth=keycloak_auth)
def parse_product_link(request, payload: LinkParseRequestSchema):
    """
    Produkt-Link parsen und Daten extrahieren
    """
    require_perm(request, 'haushalte.edit')
    parser = LinkParserService()
    data = parser.parse_url(payload.url)
    return data
